import openpyxl

def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def getCloumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_clumn)

def readData(file,sheetname,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnno).value=data
    workbook.save(file)